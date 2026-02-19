#!/usr/bin/env python3
"""
Comprehensive ingest of ALL NH DOE iPlatform data (252 sheets) into the EdOpt chatbot database.
Replaces the original ingest_iplatform.py which only handled 10 sheets.

Usage: python3 ingest_iplatform_full.py
"""
import json
import logging
import re
from datetime import datetime, timezone

import openpyxl

from models import init_db, SessionLocal, EducationStatistic

try:
    from ingest import generate_all_embeddings
    HAS_EMBEDDINGS = True
except ImportError:
    HAS_EMBEDDINGS = False

XLSX_PATH = "data/iPlatform20251207-clean.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("ingest_iplatform_full")


# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def _int(val):
    """Safely convert to int, handling None, strings with commas/dollars."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        v = int(val)
        return v if v != 0 else None
    s = str(val).strip().replace(",", "").replace("$", "").replace("--", "")
    try:
        v = int(float(s))
        return v if v != 0 else None
    except (ValueError, TypeError):
        return None


def _float(val):
    """Safely convert to float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "").replace("%", "").replace("--", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _str(val):
    """Safely convert to stripped string."""
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s != "--" else None


def _pct(val):
    """Convert a value that might be a fraction (0.87) or percent (87.5) to a display %."""
    f = _float(val)
    if f is None:
        return None
    if 0 < f < 1:
        return round(f * 100, 1)
    return round(f, 1)


def _rec(**kwargs):
    """Create an EducationStatistic record, filtering None name fields."""
    return EducationStatistic(**kwargs)


# ============================================================================
# ENROLLMENT PARSERS
# ============================================================================

def parse_district_enrollment(ws, year):
    """District Fall Enrollments. Header row 12, data from row 14."""
    records = []
    for row in ws.iter_rows(min_row=14, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        total = _int(vals[10])
        if total is None:
            continue
        records.append(_rec(
            stat_type="district_enrollment", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({
                "preschool": _int(vals[4]), "kindergarten": _int(vals[5]),
                "elementary": _int(vals[6]), "middle": _int(vals[7]),
                "high": _int(vals[8]), "pg": _int(vals[9]), "total": total,
            }),
        ))
    return records


def parse_sau_enrollment(ws, year):
    """SAU Enrollment. Header row 10, data from row 12."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        total = _int(vals[10])
        if total is None:
            continue
        records.append(_rec(
            stat_type="sau_enrollment", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({
                "preschool": _int(vals[4]), "kindergarten": _int(vals[5]),
                "elementary": _int(vals[6]), "middle": _int(vals[7]),
                "high": _int(vals[8]), "pg": _int(vals[9]), "total": total,
            }),
        ))
    return records


def parse_school_enrollment(ws, year):
    """School Enrollments by Grade. Header row 10, data from row 12."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        school_name = _str(vals[5])
        if not school_name:
            continue
        total = _int(vals[21])
        if total is None:
            continue
        data = {"preschool": _int(vals[6]), "kindergarten": _int(vals[7])}
        for g in range(1, 13):
            data[f"grade{g}"] = _int(vals[7 + g])
        data["pg"] = _int(vals[20])
        data["total"] = total
        records.append(_rec(
            stat_type="school_enrollment", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=_str(vals[3]),
            school_number=_int(vals[4]), school_name=school_name,
            data_json=json.dumps(data),
        ))
    return records


def parse_home_education(ws, year):
    """Home Education Enrollments By District. Header row 11, data from row 12."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        total = _int(vals[4])
        if total is None:
            continue
        records.append(_rec(
            stat_type="home_education", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({"total": total}),
        ))
    return records


def parse_nonpublic_enrollment(ws, year):
    """NonPublic School Enrollments. Header row 10, data from row 12."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        school_name = _str(vals[1])
        if not school_name:
            continue
        total = _int(vals[20])
        if total is None:
            continue
        data = {"ps": _int(vals[3]), "kg": _int(vals[4])}
        for g in range(1, 13):
            data[f"grade{g}"] = _int(vals[4 + g])
        data["pg"] = _int(vals[17])
        data["total"] = total
        records.append(_rec(
            stat_type="nonpublic_enrollment", school_year=year,
            school_number=_int(vals[0]), school_name=school_name,
            town=_str(vals[2]),
            data_json=json.dumps(data),
        ))
    return records


def parse_county_enrollment(ws, year):
    """County Enrollments by Grade. Row 10 headers, row 11+ data."""
    records = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        vals = list(row)
        county = _str(vals[0])
        if not county or county == "State Total":
            continue
        total = _int(vals[17])
        if total is None:
            continue
        data = {
            "preschool": _int(vals[2]), "kindergarten": _int(vals[3]),
        }
        for g in range(1, 13):
            data[f"grade{g}"] = _int(vals[3 + g])
        data["pg_sped"] = _int(vals[16])
        data["total"] = total
        records.append(_rec(
            stat_type="county_enrollment", school_year=year,
            district_name=county,
            data_json=json.dumps(data),
        ))
    return records


def parse_town_enrollment(ws, year):
    """Town Level Enrollment By Grade. Row 10 headers, row 12+ data (skip row 11 state total)."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        town_name = _str(vals[1])
        if not town_name or town_name == "Total":
            continue
        total = _int(vals[17])
        if total is None:
            continue
        data = {"preschool": _int(vals[2]), "kindergarten": _int(vals[3])}
        for g in range(1, 13):
            data[f"grade{g}"] = _int(vals[3 + g])
        data["pg"] = _int(vals[16])
        data["total"] = total
        records.append(_rec(
            stat_type="town_enrollment", school_year=year,
            town=town_name,
            data_json=json.dumps(data),
        ))
    return records


def parse_kindergarten(ws, year):
    """Kindergarten Enrollments. Row 11 headers, row 13+ data (skip row 12 state total)."""
    records = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        vals = list(row)
        school_name = _str(vals[5])
        if not school_name:
            continue
        count = _int(vals[6])
        if count is None:
            continue
        records.append(_rec(
            stat_type="kindergarten_enrollment", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=_str(vals[3]),
            school_number=_int(vals[4]), school_name=school_name,
            data_json=json.dumps({"total": count}),
        ))
    return records


def parse_preschool(ws, year):
    """Preschool Enrollments. Row 10 headers, row 13+ data (skip rows 11-12 totals)."""
    records = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        vals = list(row)
        school_name = _str(vals[5])
        if not school_name:
            continue
        count = _int(vals[6])
        if count is None:
            continue
        records.append(_rec(
            stat_type="preschool_enrollment", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=_str(vals[3]),
            school_number=_int(vals[4]), school_name=school_name,
            data_json=json.dumps({"total": count}),
        ))
    return records


def parse_hs_enrollment(ws, year):
    """High School Enrollments. Special format — grouped by size range, parse school+enrollment."""
    records = []
    current_type = None
    for row in ws.iter_rows(min_row=13, values_only=True):
        vals = list(row)
        # Check for school type header
        if _str(vals[0]) and _str(vals[1]) and "Pupils" in str(vals[1]):
            current_type = _str(vals[0])
            continue
        # Check for school type in col 0 alone
        if _str(vals[0]) and not _str(vals[1]):
            current_type = _str(vals[0])
            continue
        school_name = _str(vals[1])
        if not school_name:
            continue
        enrollment = _int(vals[2])
        if enrollment is None:
            continue
        records.append(_rec(
            stat_type="hs_enrollment", school_year=year,
            school_name=school_name,
            data_json=json.dumps({
                "total": enrollment,
                "school_type": current_type,
            }),
        ))
    return records


# ============================================================================
# FINANCIAL PARSERS
# ============================================================================

def parse_cost_per_pupil(ws, year):
    """Cost Per Pupil by district. Header row 19, data from row 23."""
    records = []
    for row in ws.iter_rows(min_row=23, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        total = _float(vals[7])
        if total is None or total == 0:
            continue
        data = {
            "elementary": round(_float(vals[4])) if _float(vals[4]) else None,
            "middle": round(_float(vals[5])) if _float(vals[5]) else None,
            "high": round(_float(vals[6])) if _float(vals[6]) else None,
            "total": round(total),
        }
        records.append(_rec(
            stat_type="cost_per_pupil", school_year=year,
            sau_number=_int(vals[2]), district_number=_int(vals[0]),
            district_name=district_name,
            data_json=json.dumps(data),
        ))
    return records


def parse_indirect_cost(ws, year):
    """Indirect Cost Rate. Header row 8, data from row 10 (skip blank row 9)."""
    records = []
    for row in ws.iter_rows(min_row=10, values_only=True):
        vals = list(row)
        name = _str(vals[3])
        if not name:
            continue
        rate = _str(vals[4])
        requested = _str(vals[5])
        if not rate:
            continue
        records.append(_rec(
            stat_type="indirect_cost_rate", school_year=year,
            sau_number=_int(vals[0]), district_number=_int(vals[1]),
            district_name=name,
            data_json=json.dumps({
                "rate": rate,
                "requested": requested,
            }),
        ))
    return records


def parse_equalized_valuation(ws, year):
    """Equalized Valuation Per Pupil — auto-detect format.
    12-col format: CTY(0), SAU(1), DIS(2), LOC(3), District(4), _, Value(6), _, ADM-R(8), _, PerPupil(10)
    10-col format: District(0), _, Value(2), _, ADM-R(4), _, PerPupil(6)"""
    # Detect format by checking column count of first data row
    sample = None
    for row in ws.iter_rows(min_row=20, max_row=25, values_only=True):
        vals = list(row)
        if _str(vals[0]) and _str(vals[0]) != "STATE AVERAGE":
            sample = vals
            break
    if sample is None:
        return []

    ncols = len(sample)
    is_old_format = ncols <= 10

    records = []
    for row in ws.iter_rows(min_row=20, values_only=True):
        vals = list(row)
        if is_old_format:
            district_name = _str(vals[0])
            if not district_name or district_name == "STATE AVERAGE":
                continue
            eq_value = _float(vals[2])
            adm_r = _float(vals[4])
            per_pupil = _float(vals[6])
        else:
            district_name = _str(vals[4])
            if not district_name or district_name == "STATE AVERAGE":
                continue
            eq_value = _float(vals[6])
            adm_r = _float(vals[8])
            per_pupil = _float(vals[10])

        if per_pupil is None:
            continue
        records.append(_rec(
            stat_type="equalized_valuation", school_year=year,
            sau_number=_int(vals[1]) if not is_old_format else None,
            district_number=_int(vals[2]) if not is_old_format else None,
            district_name=district_name,
            data_json=json.dumps({
                "equalized_value": round(eq_value) if eq_value else None,
                "adm_r": round(adm_r, 2) if adm_r else None,
                "per_pupil": round(per_pupil) if per_pupil else None,
            }),
        ))
    return records


def parse_estimated_expenditures(ws, year):
    """Estimated Expenditures — state-level summary. Rows are categories, cols are Elem/Middle/High/Total."""
    records = []
    for row in ws.iter_rows(min_row=10, values_only=True):
        vals = list(row)
        category = _str(vals[0])
        if not category:
            continue
        total = _float(vals[8]) if len(vals) > 8 else _float(vals[7]) if len(vals) > 7 else None
        if total is None:
            continue
        records.append(_rec(
            stat_type="estimated_expenditures", school_year=year,
            district_name=category,
            data_json=json.dumps({
                "elementary": _float(vals[2]) if len(vals) > 2 else None,
                "middle": _float(vals[4]) if len(vals) > 4 else None,
                "high": _float(vals[6]) if len(vals) > 6 else None,
                "total": total,
            }),
        ))
    return records


def parse_state_avg_cpp(ws, year):
    """State Average Cost Per Pupil — state-level summary."""
    records = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        vals = list(row)
        category = _str(vals[0])
        if not category:
            continue
        total = _float(vals[7]) if len(vals) > 7 else _float(vals[6]) if len(vals) > 6 else None
        if total is None:
            continue
        records.append(_rec(
            stat_type="state_avg_cpp", school_year=year,
            district_name=category,
            data_json=json.dumps({
                "elementary": _float(vals[1]),
                "middle": _float(vals[3]) if len(vals) > 3 else None,
                "high": _float(vals[5]) if len(vals) > 5 else None,
                "total": total,
            }),
        ))
    return records


def parse_summary_rev_exp(ws, year):
    """State Summary Revenue and Expenditures."""
    records = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        vals = list(row)
        category = _str(vals[0])
        if not category:
            continue
        amount = _float(vals[1])
        pct = _pct(vals[3]) if len(vals) > 3 else None
        if amount is None:
            continue
        records.append(_rec(
            stat_type="state_revenue_expenditure", school_year=year,
            district_name=category,
            data_json=json.dumps({
                "amount": round(amount),
                "pct": pct,
            }),
        ))
    return records


def parse_adm(ws, year):
    """Average Daily Membership. Row 11 headers, row 13+ data.
    ADM-A in cols 2-8, ADM-R in cols 10-17."""
    records = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        vals = list(row)
        district_name = _str(vals[1])
        if not district_name or district_name == "State Totals":
            continue
        adm_a_total = _float(vals[8])
        adm_r_total = _float(vals[16])
        if adm_a_total is None and adm_r_total is None:
            continue
        records.append(_rec(
            stat_type="adm", school_year=year,
            district_number=_int(vals[0]), district_name=district_name,
            data_json=json.dumps({
                "adm_a_preschool": _float(vals[2]),
                "adm_a_kindergarten": _float(vals[3]),
                "adm_a_elementary": _float(vals[4]),
                "adm_a_middle": _float(vals[5]),
                "adm_a_total_elem": _float(vals[6]),
                "adm_a_high": _float(vals[7]),
                "adm_a_total": round(adm_a_total, 2) if adm_a_total else None,
                "adm_r_preschool": _float(vals[10]),
                "adm_r_kindergarten": _float(vals[11]),
                "adm_r_elementary": _float(vals[12]),
                "adm_r_middle": _float(vals[13]),
                "adm_r_total_elem": _float(vals[14]),
                "adm_r_high": _float(vals[15]),
                "adm_r_total": round(adm_r_total, 2) if adm_r_total else None,
            }),
        ))
    return records


# ============================================================================
# DEMOGRAPHICS PARSERS
# ============================================================================

def parse_free_reduced_lunch(ws, year):
    """Free Reduced K-12 School Lunch. Header row 10, data from row 18."""
    records = []
    for row in ws.iter_rows(min_row=18, values_only=True):
        vals = list(row)
        school_name = _str(vals[5])
        district_name = _str(vals[3])
        if not school_name and not district_name:
            continue
        if not _int(vals[0]) and not _int(vals[4]):
            continue
        enrollment = _int(vals[6])
        eligible = _int(vals[7])
        pct = _float(vals[8])
        if enrollment is None:
            continue
        data = {
            "enrollment": enrollment,
            "eligible": eligible or 0,
            "pct_eligible": round(pct * 100, 1) if pct is not None and pct < 1 else pct,
        }
        records.append(_rec(
            stat_type="free_reduced_lunch", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            school_number=_int(vals[4]), school_name=school_name,
            data_json=json.dumps(data),
        ))
    return records


def parse_free_reduced_preschool(ws, year):
    """Free Reduced PRE-School Lunch. Same format as K-12 version."""
    return parse_free_reduced_lunch(ws, year)


def parse_lep(ws, year):
    """Limited English Proficiency. Row 10 headers, row 12+ data."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        enrollment = _int(vals[4])
        if enrollment is None:
            continue
        records.append(_rec(
            stat_type="limited_english", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({
                "enrollment": enrollment,
                "el_eligible": _int(vals[5]),
                "el_monitor": _int(vals[6]),
            }),
        ))
    return records


def parse_race_ethnic_district(ws, year):
    """Race/Ethnic Enrollments by District. Row 11-12 headers (2-row), row 14+ data.
    Cols: Dist#(0), DistName(1), then pairs of Count/% for each race."""
    records = []
    for row in ws.iter_rows(min_row=14, values_only=True):
        vals = list(row)
        district_name = _str(vals[1])
        if not district_name or district_name == "State Total":
            continue
        total = _int(vals[15])
        if total is None:
            continue
        records.append(_rec(
            stat_type="race_ethnic", school_year=year,
            district_number=_int(vals[0]), district_name=district_name,
            data_json=json.dumps({
                "american_indian": _int(vals[2]),
                "american_indian_pct": _str(vals[3]),
                "asian_pacific": _int(vals[4]),
                "asian_pacific_pct": _str(vals[5]),
                "hispanic": _int(vals[6]),
                "hispanic_pct": _str(vals[7]),
                "black": _int(vals[8]),
                "black_pct": _str(vals[9]),
                "white": _int(vals[10]),
                "white_pct": _str(vals[11]),
                "multi_race": _int(vals[12]),
                "multi_race_pct": _str(vals[13]),
                "total": total,
            }),
        ))
    return records


def parse_race_ethnic_school(ws, year):
    """Race/Ethnic Enrollments by School (K-12). Row 11-12 headers, row 14+ data.
    Cols: SAU(0), SAU Name(1), Dist(2), Dist Name(3), SchID(4), School(5),
          then pairs of F&R/Enroll for each race category."""
    records = []
    for row in ws.iter_rows(min_row=14, values_only=True):
        vals = list(row)
        school_name = _str(vals[5])
        if not school_name or school_name == "State Total":
            continue
        # Total is sum of all enrollment columns
        # White enroll is at col 20
        white_enroll = _int(vals[20]) if len(vals) > 20 else None
        if white_enroll is None and not _int(vals[7]):
            continue
        records.append(_rec(
            stat_type="race_ethnic_school", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=_str(vals[3]),
            school_number=_int(vals[4]), school_name=school_name,
            data_json=json.dumps({
                "am_indian_fr": _int(vals[6]), "am_indian_enroll": _int(vals[7]),
                "asian_fr": _int(vals[8]), "asian_enroll": _int(vals[9]),
                "black_fr": _int(vals[11]), "black_enroll": _int(vals[12]),
                "hispanic_fr": _int(vals[13]), "hispanic_enroll": _int(vals[14]),
                "multi_race_fr": _int(vals[15]), "multi_race_enroll": _int(vals[16]),
                "nh_pacific_fr": _int(vals[17]), "nh_pacific_enroll": _int(vals[18]),
                "white_fr": _int(vals[19]), "white_enroll": white_enroll,
            }),
        ))
    return records


# ============================================================================
# OUTCOMES PARSERS
# ============================================================================

def parse_attendance(ws, year):
    """Attendance Rate By District. Row 10 headers, row 12+ data."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        total_pct = _float(vals[9])
        if total_pct is None:
            continue
        records.append(_rec(
            stat_type="attendance_rate", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({
                "preschool_pct": _float(vals[4]),
                "kindergarten_pct": _float(vals[5]),
                "elementary_pct": _float(vals[6]),
                "middle_pct": _float(vals[7]),
                "high_pct": _float(vals[8]),
                "total_pct": round(total_pct, 1),
            }),
        ))
    return records


def parse_cohort(ws, year):
    """Cohort Graduation/Dropout Rate By School. Row 12 has SAU/Dist/School labels, row 13+ data."""
    records = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        vals = list(row)
        school_name = _str(vals[3])
        if not school_name or school_name == "State Total":
            continue
        cohort_size = _int(vals[4])
        if cohort_size is None:
            continue
        records.append(_rec(
            stat_type="cohort_graduation", school_year=year,
            sau_number=_int(vals[0]), district_number=_int(vals[1]),
            district_name=_str(vals[2]), school_name=school_name,
            data_json=json.dumps({
                "cohort_size": cohort_size,
                "graduated": _int(vals[5]),
                "graduation_rate": _pct(vals[6]),
                "hiset": _int(vals[7]),
                "dropped_out": _int(vals[8]),
                "total_non_graduates": _int(vals[9]),
                "dropout_rate": _pct(vals[10]),
            }),
        ))
    return records


def parse_completers_category(ws, year):
    """State Completers By Category. Row 10 headers, row 12+ data."""
    records = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        vals = list(row)
        category = _str(vals[0])
        if not category:
            continue
        count = _int(vals[1])
        if count is None:
            continue
        records.append(_rec(
            stat_type="completers_category", school_year=year,
            district_name=category,
            data_json=json.dumps({
                "count": count,
                "male_pct": _pct(vals[2]),
                "female_pct": _pct(vals[3]),
                "total_pct": _pct(vals[4]),
            }),
        ))
    return records


def parse_completers_county(ws, year):
    """Completers By Status By County. Row 11 headers, row 13+ data."""
    records = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        vals = list(row)
        county = _str(vals[0])
        if not county or county == "Total":
            continue
        total = _int(vals[1])
        if total is None:
            continue
        records.append(_rec(
            stat_type="completers_county", school_year=year,
            district_name=county,
            data_json=json.dumps({
                "total": total,
                "four_year_college_pct": _pct(vals[2]),
                "less_than_four_year_pct": _pct(vals[3]),
                "post_grad_pct": _pct(vals[4]),
                "employed_pct": _pct(vals[5]),
                "armed_forces_pct": _pct(vals[6]),
                "unemployed_pct": _pct(vals[7]),
                "unknown_pct": _pct(vals[8]),
            }),
        ))
    return records


def parse_completers_school(ws, year):
    """Completers By Status By School. Row 11 headers, row 13+ data."""
    records = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        vals = list(row)
        school_name = _str(vals[5])
        if not school_name or school_name == "State Total":
            continue
        total = _int(vals[6])
        if total is None:
            continue
        records.append(_rec(
            stat_type="completers_school", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=_str(vals[3]),
            school_number=_int(vals[4]), school_name=school_name,
            data_json=json.dumps({
                "total": total,
                "four_year_college_pct": _pct(vals[7]),
                "less_than_four_year_pct": _pct(vals[8]),
                "post_grad_pct": _pct(vals[9]),
                "employed_pct": _pct(vals[10]),
                "armed_forces_pct": _pct(vals[11]),
                "unemployed_pct": _pct(vals[12]),
                "unknown_pct": _pct(vals[13]),
            }),
        ))
    return records


# ============================================================================
# STAFFING PARSERS
# ============================================================================

def parse_class_size_district(ws, year):
    """Average Class Size By District. Row 12 headers, row 14+ data (skip row 13 state avg)."""
    records = []
    for row in ws.iter_rows(min_row=14, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        records.append(_rec(
            stat_type="avg_class_size", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({
                "grades_1_2": _float(vals[4]),
                "grades_3_4": _float(vals[5]),
                "grades_5_8": _float(vals[6]),
            }),
        ))
    return records


def parse_class_size_school(ws, year):
    """Average Class Size By School. Row 11 headers, row 13+ data (skip row 12 state avg)."""
    records = []
    for row in ws.iter_rows(min_row=13, values_only=True):
        vals = list(row)
        school_name = _str(vals[4])
        if not school_name or _str(vals[3]) == "State Average":
            continue
        records.append(_rec(
            stat_type="avg_class_size_school", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=_str(vals[3]),
            school_name=school_name,
            data_json=json.dumps({
                "grades_1_2": _float(vals[5]),
                "grades_3_4": _float(vals[6]),
                "grades_5_8": _float(vals[7]),
            }),
        ))
    return records


def parse_staff_fte(ws, year):
    """Staff FTE by District. Row 10 headers, row 12+ data."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        records.append(_rec(
            stat_type="staff_fte", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({
                "teachers": _float(vals[4]),
                "instruction_support": _float(vals[5]),
                "librarians": _float(vals[6]),
                "specialists": _float(vals[7]),
                "admin_support": _float(vals[8]),
                "all_other": _float(vals[9]),
            }),
        ))
    return records


def parse_student_teacher_ratio(ws, year):
    """Student Teacher Ratio by District. Row 10 headers, row 12+ data."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        records.append(_rec(
            stat_type="student_teacher_ratio", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({
                "enrollment": _int(vals[4]),
                "teachers": _float(vals[5]),
                "ratio": _float(vals[6]),
                "prev_enrollment": _int(vals[7]),
                "prev_teachers": _float(vals[8]) if _float(vals[8]) else _str(vals[8]),
                "prev_ratio": _float(vals[9]),
            }),
        ))
    return records


def parse_teacher_attainment(ws, year):
    """Teacher Attainment by District. Row 10 headers, row 12+ data."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        records.append(_rec(
            stat_type="teacher_attainment", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({
                "num_teachers": _float(vals[4]),
                "pct_bachelors": _float(vals[5]),
                "pct_masters": _float(vals[6]),
                "pct_beyond_masters": _float(vals[7]),
            }),
        ))
    return records


def parse_teacher_salary(ws, year):
    """Teacher Average Salary by District. Row 10 headers, row 12+ data."""
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        salary = _float(vals[5])
        if salary is None:
            continue
        records.append(_rec(
            stat_type="teacher_salary", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({
                "num_teachers": _float(vals[4]),
                "avg_salary": round(salary),
            }),
        ))
    return records


def parse_teacher_salary_early(ws, year):
    """Older teacher salary format (pre-2022). Auto-detects column layout.
    Formats: 6-col (SAU/District), 4-col (None/District/Teachers/Salary),
             3-col (District/Teachers/Salary)."""
    records = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        vals = list(row)
        ncols = len(vals)

        # 4-col format: None(0), District(1), NumTeachers(2), Salary(3)
        if ncols == 4:
            district_name = _str(vals[1])
            if not district_name or "average" in district_name.lower():
                continue
            salary = _float(vals[3])
            if salary is None or salary < 10000:
                continue
            records.append(_rec(
                stat_type="teacher_salary", school_year=year,
                district_name=district_name,
                data_json=json.dumps({
                    "num_teachers": _float(vals[2]),
                    "avg_salary": round(salary),
                }),
            ))
            continue

        # 3-col format: District(0), NumTeachers(1), Salary(2)
        if ncols == 3:
            district_name = _str(vals[0])
            if not district_name or "average" in district_name.lower() or "district" in district_name.lower():
                continue
            salary = _float(vals[2])
            if salary is None or salary < 10000:
                continue
            records.append(_rec(
                stat_type="teacher_salary", school_year=year,
                district_name=district_name,
                data_json=json.dumps({
                    "num_teachers": _float(vals[1]),
                    "avg_salary": round(salary),
                }),
            ))
            continue

        # 6-col format: SAU#(0), SAUName(1), Dist#(2), DistName(3), Teachers(4), Salary(5)
        if ncols >= 6:
            district_name = _str(vals[3])
            if not district_name or "average" in district_name.lower():
                continue
            salary = _float(vals[5])
            if salary is None or salary < 10000:
                continue
            records.append(_rec(
                stat_type="teacher_salary", school_year=year,
                sau_number=_int(vals[0]), sau_name=_str(vals[1]),
                district_number=_int(vals[2]), district_name=district_name,
                data_json=json.dumps({
                    "num_teachers": _float(vals[4]),
                    "avg_salary": round(salary),
                }),
            ))
    return records


def parse_teacher_starting_salary(ws, year):
    """Teacher Minimum Starting Salary — auto-detects format.
    6-col: Ranking(0), SAU#(1), SAUName(2), Dist#(3), DistName(4), Salary(5)
    7-col side-by-side: Ranking(0), District(1), Salary(2), None(3), Ranking(4), District(5), Salary(6)"""
    # Detect format from first data-like row
    ncols = 0
    for row in ws.iter_rows(min_row=6, max_row=20, values_only=True):
        vals = list(row)
        ncols = len(vals)
        if ncols >= 6 and _int(vals[0]) is not None:
            break

    records = []
    seen = set()  # Deduplicate districts in side-by-side format

    for row in ws.iter_rows(min_row=6, values_only=True):
        vals = list(row)

        if ncols >= 7 and (len(vals) < 6 or (len(vals) >= 6 and _str(vals[3]) is None)):
            # 7-col side-by-side format: process both left and right columns
            for offset in (0, 4):
                if offset + 2 >= len(vals):
                    continue
                district_name = _str(vals[offset + 1])
                if not district_name or "ranking" in district_name.lower() or "average" in district_name.lower():
                    continue
                salary = _float(vals[offset + 2])
                if salary is None or salary < 10000:
                    continue
                if district_name in seen:
                    continue
                seen.add(district_name)
                records.append(_rec(
                    stat_type="teacher_salary", school_year=year,
                    district_name=district_name,
                    data_json=json.dumps({
                        "min_starting_salary": round(salary),
                        "salary_ranking": _int(vals[offset]),
                    }),
                ))
        else:
            # 6-col SAU format
            if len(vals) < 6:
                continue
            district_name = _str(vals[4])
            if not district_name or "average" in district_name.lower() or "ranking" in str(vals[0]).lower():
                continue
            salary = _float(vals[5])
            if salary is None or salary < 10000:
                continue
            records.append(_rec(
                stat_type="teacher_salary", school_year=year,
                sau_number=_int(vals[1]), sau_name=_str(vals[2]),
                district_number=_int(vals[3]), district_name=district_name,
                data_json=json.dumps({
                    "min_starting_salary": round(salary),
                    "salary_ranking": _int(vals[0]),
                }),
            ))
    return records


def parse_teacher_salary_schedule(ws, year):
    """Teacher Salary Schedule. Row 10 headers, row 11+ data."""
    records = []
    for row in ws.iter_rows(min_row=11, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        degree = _str(vals[4])
        min_sal = _float(vals[5])
        max_sal = _float(vals[6])
        steps = _int(vals[7])
        if not degree:
            continue
        records.append(_rec(
            stat_type="teacher_salary_schedule", school_year=year,
            sau_number=_int(vals[0]), sau_name=_str(vals[1]),
            district_number=_int(vals[2]), district_name=district_name,
            data_json=json.dumps({
                "degree_type": degree,
                "min_salary": round(min_sal) if min_sal else None,
                "max_salary": round(max_sal) if max_sal else None,
                "steps": steps,
            }),
        ))
    return records


def parse_principal_salary(ws, year):
    """Principal Salary — dynamic format detection.
    Scans for header row containing 'Salary', then maps columns by name."""
    # Find header row (scan rows 1-15 for a row containing "Salary")
    header_row_num = None
    headers = []
    for r in range(1, 16):
        try:
            row_data = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        except IndexError:
            continue
        row_strs = [str(c).lower() if c else "" for c in row_data]
        if any("salary" in s for s in row_strs) and any("school" in s or "sch " in s or "sch id" in s for s in row_strs):
            header_row_num = r
            headers = row_strs
            break

    if header_row_num is None:
        logger.warning(f"Could not find header row for principal salary {year}")
        return []

    # Map column indices
    def find_col(*keywords):
        for i, h in enumerate(headers):
            for kw in keywords:
                if kw in h:
                    return i
        return None

    i_sau = find_col("sau id", "sau #")
    i_sau_name = find_col("sau name", "sau ")
    i_dst = find_col("dst id", "dst #", "dis id", "dis #", "district id")
    i_dst_name = find_col("dst name", "dist name", "district name")
    i_sch_id = find_col("sch id", "sch #", "school id")
    i_sch_name = find_col("school name", "school", "sch name")
    i_salary = find_col("salary")
    i_title = find_col("contact type", "title", "position")
    i_fulltime = find_col("fulltime", "full time", "full_time")

    if i_sch_name is None or i_salary is None:
        logger.warning(f"Missing school/salary columns in principal salary {year}. Headers: {headers}")
        return []

    records = []
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        vals = list(row)
        school_name = _str(vals[i_sch_name]) if i_sch_name < len(vals) else None
        if not school_name:
            continue
        salary = _float(vals[i_salary]) if i_salary < len(vals) else None
        if salary is None:
            continue
        data = {
            "contact_type": _str(vals[i_title]) if i_title is not None and i_title < len(vals) else None,
            "salary": round(salary),
        }
        if i_fulltime is not None and i_fulltime < len(vals):
            data["full_time"] = _str(vals[i_fulltime])
        records.append(_rec(
            stat_type="principal_salary", school_year=year,
            sau_number=_int(vals[i_sau]) if i_sau is not None and i_sau < len(vals) else None,
            sau_name=_str(vals[i_sau_name]) if i_sau_name is not None and i_sau_name < len(vals) else None,
            district_number=_int(vals[i_dst]) if i_dst is not None and i_dst < len(vals) else None,
            district_name=_str(vals[i_dst_name]) if i_dst_name is not None and i_dst_name < len(vals) else None,
            school_number=_int(vals[i_sch_id]) if i_sch_id is not None and i_sch_id < len(vals) else None,
            school_name=school_name,
            data_json=json.dumps(data),
        ))
    return records


def parse_admin_salary(ws, year):
    """Admin Salary by SAU. Row 9 headers, row 10+ data."""
    records = []
    for row in ws.iter_rows(min_row=10, values_only=True):
        vals = list(row)
        sau_name = _str(vals[1])
        if not sau_name:
            continue
        salary = _float(vals[3])
        if salary is None:
            continue
        records.append(_rec(
            stat_type="admin_salary", school_year=year,
            sau_number=_int(vals[0]), sau_name=sau_name,
            data_json=json.dumps({
                "contact_type": _str(vals[2]),
                "salary": round(salary),
            }),
        ))
    return records


def parse_admin_salary_by_name(ws, year):
    """Admin Salary listed by individual name (2018-19 format).
    3 cols: Name(0), Unit(1), Salary(2). Organized by role sections (Superintendents, etc.)."""
    records = []
    current_role = None
    for row in ws.iter_rows(min_row=3, values_only=True):
        vals = list(row)
        name = _str(vals[0])
        if not name:
            continue
        # Check for section headers
        if name in ("Superintendents", "Assistant Superintendents", "Business Administrators",
                     "Business Administrator", "Superintendent", "Assistant Superintendent"):
            current_role = name.rstrip("s")  # Normalize to singular
            continue
        salary = _float(vals[2]) if len(vals) > 2 else None
        if salary is None or salary < 10000:
            continue
        unit = _str(vals[1])
        # Extract SAU number from unit string like "   17-Sanborn Regional"
        sau_num = None
        sau_name = unit
        if unit:
            m = re.match(r'\s*(\d+)\s*-\s*(.*)', unit)
            if m:
                sau_num = int(m.group(1))
                sau_name = m.group(2).strip()
        records.append(_rec(
            stat_type="admin_salary", school_year=year,
            sau_number=sau_num, sau_name=sau_name,
            data_json=json.dumps({
                "contact_type": current_role or "Administrator",
                "salary": round(salary),
                "name": name,
            }),
        ))
    return records


# ============================================================================
# TREND / SUMMARY PARSERS
# ============================================================================

def parse_home_ed_trend(ws, year_label):
    """Home Education Fall Enrollments — 10-year trend by grade.
    Row 11: year columns. Row 12+: grade rows with counts."""
    # Read year headers from row 11
    header_row = list(ws.iter_rows(min_row=11, max_row=11, values_only=True))[0]
    years = [_str(c) for c in header_row[2:]]  # Skip first 2 cols (None, None)

    records = []
    section = "public"  # Track Public District vs Non-Public
    for row in ws.iter_rows(min_row=10, values_only=True):
        vals = list(row)
        # Check for section headers
        label = _str(vals[1]) if len(vals) > 1 else None
        if label and "Non-Public" in str(label):
            section = "nonpublic"
            continue
        if label and "Grand Total" in str(label):
            section = "total"
            continue
        if label and "Public District" in str(label):
            section = "public"
            continue

        grade = _str(vals[1])
        if not grade or grade in ("Total", "Grand Total") or not _int(vals[2]):
            # Check if it's a total row
            if grade and ("Total" in grade):
                grade = f"{section}_total"
            else:
                continue

        year_data = {}
        for i, yr in enumerate(years):
            if yr and i + 2 < len(vals):
                year_data[yr] = _int(vals[i + 2])

        if year_data:
            records.append(_rec(
                stat_type="home_ed_trend", school_year=year_label,
                district_name=f"{section}: {grade}",
                data_json=json.dumps(year_data),
            ))
    return records


def parse_state_totals(ws, year_label):
    """State Totals Ten Years — enrollment trend. Same structure as home ed trend."""
    header_row = list(ws.iter_rows(min_row=11, max_row=11, values_only=True))[0]
    years = [_str(c) for c in header_row[1:]]

    records = []
    section = "public"
    for row in ws.iter_rows(min_row=10, values_only=True):
        vals = list(row)
        label = _str(vals[0])
        if label and "Non-Public" in str(label):
            section = "nonpublic"
            continue
        if label and "Grand Total" in str(label):
            section = "total"
            continue
        if label and "Public District" in str(label):
            section = "public"
            continue
        if label and "Charter" in str(label):
            section = "charter"
            continue

        grade = _str(vals[0])
        if not grade:
            continue

        year_data = {}
        for i, yr in enumerate(years):
            if yr and i + 1 < len(vals):
                year_data[yr] = _int(vals[i + 1])

        if any(v for v in year_data.values() if v):
            records.append(_rec(
                stat_type="state_totals", school_year=year_label,
                district_name=f"{section}: {grade}",
                data_json=json.dumps(year_data),
            ))
    return records


# ============================================================================
# ASSESSMENT PARSERS
# ============================================================================

SUBJECT_NAMES = {"mat": "Math", "rea": "Reading/ELA", "sci": "Science"}


def parse_assessment_2022_format(ws, school_year):
    """Parse assessment minimal sheets with yearid-first column order (2018-2022).
    Cols: yearid(0), Level(1), Subject(2), DenominatorType(3),
          District(4), Discode(5), School(6), Schcode(7), Grade(8),
          NumberStudents(9), plevel1-4(10-13), pAboveprof(14), pBelowProf(15), AvgScore(16)"""
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        school_name = _str(vals[6])
        if not school_name:
            continue
        grade_raw = vals[8]
        grade = "all" if grade_raw in (0, "0") else grade_raw
        records.append(_rec(
            stat_type="assessment", school_year=school_year,
            district_name=_str(vals[4]),
            school_number=_int(vals[7]), school_name=school_name,
            data_json=json.dumps({
                "subject": _str(vals[2]),
                "subject_name": SUBJECT_NAMES.get(_str(vals[2]), _str(vals[2])),
                "grade": grade,
                "num_students": _str(vals[9]),
                "pct_above_proficient": _str(vals[14]),
                "pct_below_proficient": _str(vals[15]),
                "avg_score": _str(vals[16]) if _str(vals[16]) and _str(vals[16]) not in ("Not available", "NULL") else None,
                "level1_pct": _str(vals[10]),
                "level2_pct": _str(vals[11]),
                "level3_pct": _str(vals[12]),
                "level4_pct": _str(vals[13]),
            }),
        ))
    return records


def parse_assessment_2023_format(ws, school_year):
    """Parse 2023 minimal assessment — same as 2022 but with DateInserted col at end.
    Cols: yearid(0), Level(1), Subject(2), DenominatorType(3),
          District(4), Discode(5), School(6), Schcode(7), Grade(8),
          NumberStudents(9), plevel1-4(10-13), pAboveprof(14), pBelowProf(15), AvgScore(16), DateInserted(17)"""
    return parse_assessment_2022_format(ws, school_year)


def parse_assessment_2024_format(ws, school_year):
    """Parse 2024 minimal assessment — different column order.
    Cols: DenominatorType(0), yearid(1), Level(2), Subject(3),
          District(4), School(5), Grade(6), Discode(7), Schcode(8),
          NumberStudents(9), plevel1-4(10-13), pAboveprof(14), pBelowProf(15), AvgScore(16), DateInserted(17)"""
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        school_name = _str(vals[5])
        if not school_name:
            continue
        grade_raw = vals[6]
        grade = "all" if grade_raw in (0, "0") else grade_raw
        records.append(_rec(
            stat_type="assessment", school_year=school_year,
            district_name=_str(vals[4]),
            school_number=_int(vals[8]), school_name=school_name,
            data_json=json.dumps({
                "subject": _str(vals[3]),
                "subject_name": SUBJECT_NAMES.get(_str(vals[3]), _str(vals[3])),
                "grade": grade,
                "num_students": _str(vals[9]),
                "pct_above_proficient": _str(vals[14]),
                "pct_below_proficient": _str(vals[15]),
                "avg_score": _str(vals[16]) if _str(vals[16]) and _str(vals[16]) not in ("Not available", "NULL") else None,
                "level1_pct": _str(vals[10]),
                "level2_pct": _str(vals[11]),
                "level3_pct": _str(vals[12]),
                "level4_pct": _str(vals[13]),
            }),
        ))
    return records


def parse_assessment_old_format(ws, school_year):
    """Parse older assessment sheets (2009-2017) — slightly different column names but same positions.
    Cols: yearid(0), replevel(1), subject(2), DenominatorType(3),
          disname(4), discode(5), schname(6), schcode(7), grade(8),
          NumberStudents(9), plevel1-4(10-13), pAboveprof(14), pBelowProf(15), AvgScore(16)"""
    return parse_assessment_2022_format(ws, school_year)


def parse_assessment_disaggregated(ws, school_year):
    """Parse large disaggregated assessment sheets. 70-90K rows.
    Different format — need to detect columns from header row."""
    headers = [_str(c) for c in list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]]

    # Build column index map
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[h.lower()] = i

    # Find key column indices
    def find_col(*names):
        for n in names:
            if n.lower() in col_map:
                return col_map[n.lower()]
        return None

    i_subject = find_col("Subject", "subject")
    i_district = find_col("District", "disname")
    i_school = find_col("School", "schname")
    i_grade = find_col("Grade", "grade")
    i_schcode = find_col("Schcode", "schcode")
    i_numstud = find_col("NumberStudents", "numberstudents")
    i_above = find_col("pAboveprof", "paboveprof")
    i_below = find_col("pBelowProf", "pbelowprof")
    i_denom = find_col("DenominatorType", "denominatortype")
    i_aggregated = find_col("Aggregated by", "aggregated by")

    if i_school is None or i_subject is None:
        logger.warning(f"Could not find required columns in disaggregated sheet. Headers: {headers[:10]}")
        return []

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        school_name = _str(vals[i_school]) if i_school is not None else None
        if not school_name:
            continue
        grade_raw = vals[i_grade] if i_grade is not None else None
        grade = "all" if grade_raw in (0, "0") else grade_raw

        data = {
            "subject": _str(vals[i_subject]),
            "grade": grade,
            "num_students": _str(vals[i_numstud]) if i_numstud is not None else None,
            "pct_above_proficient": _str(vals[i_above]) if i_above is not None else None,
            "pct_below_proficient": _str(vals[i_below]) if i_below is not None else None,
        }
        if i_denom is not None:
            data["denominator_type"] = _str(vals[i_denom])
        if i_aggregated is not None:
            data["aggregated_by"] = _str(vals[i_aggregated])

        records.append(_rec(
            stat_type="assessment_disaggregated", school_year=school_year,
            district_name=_str(vals[i_district]) if i_district is not None else None,
            school_number=_int(vals[i_schcode]) if i_schcode is not None else None,
            school_name=school_name,
            data_json=json.dumps(data),
        ))
    return records


def parse_assessment_by_town(ws, school_year):
    """Parse assessment by town of responsibility sheets. Same approach as disaggregated."""
    return parse_assessment_disaggregated(ws, school_year)


# ============================================================================
# PROPERTY TAX / VALUES PARSER
# ============================================================================

def parse_values_tax(ws, year):
    """Property Tax Values — very wide sheets (82+ cols). Extract key columns only.
    Multi-row header: scan for a row containing 'DISTRICT' in any cell."""
    # Find the row containing "DISTRICT" as a cell value
    header_row = None
    dist_col = None
    for r in range(1, 30):
        try:
            row_data = list(ws.iter_rows(min_row=r, max_row=r, values_only=True))[0]
        except IndexError:
            continue
        for i, c in enumerate(row_data):
            if c and str(c).strip().upper() == "DISTRICT":
                header_row = r
                dist_col = i
                break
        if header_row:
            break

    if header_row is None:
        logger.warning(f"Could not find DISTRICT column in values/tax sheet for {year}")
        return []

    headers = [_str(c) for c in list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]]

    records = []
    # Data typically starts 2-3 rows after header (skip blank row)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        vals = list(row)
        district = _str(vals[dist_col]) if dist_col < len(vals) else None
        if not district:
            continue
        # Store all numeric values from non-district columns
        data = {}
        for i, h in enumerate(headers):
            if h and i < len(vals) and vals[i] is not None and i != dist_col:
                val = _float(vals[i])
                if val is not None:
                    data[h] = val
        if data:
            records.append(_rec(
                stat_type="property_tax_values", school_year=year,
                district_name=district,
                data_json=json.dumps(data),
            ))
    return records


# ============================================================================
# STATE PROGRAMS
# ============================================================================

def parse_state_programs(ws, year):
    """US DOE formula-allocated funds by program."""
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        if len(vals) < 4:
            continue
        program = _str(vals[0]) or _str(vals[1])
        if not program:
            continue
        amount = _float(vals[3]) if len(vals) > 3 else _float(vals[2])
        if amount is None:
            continue
        records.append(_rec(
            stat_type="state_programs", school_year=year,
            district_name=program,
            data_json=json.dumps({"amount": amount}),
        ))
    return records


# ============================================================================
# COMPREHENSIVE SHEET MAPPING
# ============================================================================

# Maps sheet_name -> (parser_function, school_year)
SHEET_MAP = {
    # === DISTRICT ENROLLMENT (current + 14 historical) ===
    "District Fall Enrollments_Distr": (parse_district_enrollment, "2025-26"),
    "District Fall Enrollments_Dis_1": (parse_district_enrollment, "2024-25"),
    "District Fall Enrollments (1)_D": (parse_district_enrollment, "2023-24"),
    "District Fall Enrollments (2)_D": (parse_district_enrollment, "2022-23"),
    "District Fall Enrollments (3)_D": (parse_district_enrollment, "2021-22"),
    "District Fall Enrollments (4)_D": (parse_district_enrollment, "2020-21"),
    "District Fall Enrollments (5)_D": (parse_district_enrollment, "2019-20"),
    "District Fall Enrollments (6)_D": (parse_district_enrollment, "2018-19"),
    "District Fall Enrollments (7)_D": (parse_district_enrollment, "2017-18"),
    "District Fall Enrollments (8)_D": (parse_district_enrollment, "2016-17"),
    "District Fall Enrollments (9)_D": (parse_district_enrollment, "2015-16"),
    "District Fall Enrollments (10)_": (parse_district_enrollment, "2014-15"),
    "District Fall Enrollments (11)_": (parse_district_enrollment, "2013-14"),
    "District Fall Enrollments (12)_": (parse_district_enrollment, "2012-13"),
    "District Fall Enrollments (13)_": (parse_district_enrollment, "2011-12"),

    # === SAU ENROLLMENT (current + 15 historical) ===
    "School Administrative Unit Enro": (parse_sau_enrollment, "2025-26"),
    "School Administrative Unit En_1": (parse_sau_enrollment, "2024-25"),
    "School Administrative Unit En_2": (parse_sau_enrollment, "2025-26"),  # duplicate per iPlatform
    "School Administrative Unit En_3": (parse_sau_enrollment, "2023-24"),
    "School Administrative Unit En_4": (parse_sau_enrollment, "2022-23"),
    "School Administrative Unit En_5": (parse_sau_enrollment, "2021-22"),
    "School Administrative Unit En_6": (parse_sau_enrollment, "2020-21"),
    "School Administrative Unit En_7": (parse_sau_enrollment, "2019-20"),
    "School Administrative Unit En_8": (parse_sau_enrollment, "2018-19"),
    "School Administrative Unit En_9": (parse_sau_enrollment, "2017-18"),
    "Recovered_Sheet6": (parse_sau_enrollment, "2016-17"),
    "Recovered_Sheet7": (parse_sau_enrollment, "2015-16"),
    "Recovered_Sheet8": (parse_sau_enrollment, "2014-15"),
    "Recovered_Sheet9": (parse_sau_enrollment, "2013-14"),
    "Recovered_Sheet10": (parse_sau_enrollment, "2012-13"),
    "Recovered_Sheet11": (parse_sau_enrollment, "2011-12"),

    # === SCHOOL ENROLLMENT ===
    "School Enrollments by Grade Pub": (parse_school_enrollment, "2025-26"),

    # === HOME EDUCATION (current + 15 historical) ===
    "Home Education Enrollments By D": (parse_home_education, "2025-26"),
    "Home Education Enrollments By_1": (parse_home_education, "2025-26"),  # duplicate
    "Home Education Enrollments By_2": (parse_home_education, "2024-25"),
    "Home Education Enrollments By_3": (parse_home_education, "2023-24"),
    "Home Education Enrollments By_4": (parse_home_education, "2022-23"),
    "Home Education Enrollments By_5": (parse_home_education, "2021-22"),
    "Home Education Enrollments By_6": (parse_home_education, "2020-21"),
    "Home Education Enrollments By_7": (parse_home_education, "2019-20"),
    "Home Education Enrollments By_8": (parse_home_education, "2018-19"),
    "Home Education Enrollments By_9": (parse_home_education, "2017-18"),
    "Recovered_Sheet12": (parse_home_education, "2016-17"),
    "Recovered_Sheet13": (parse_home_education, "2015-16"),
    "Recovered_Sheet14": (parse_home_education, "2014-15"),
    "Recovered_Sheet15": (parse_home_education, "2013-14"),
    "Recovered_Sheet16": (parse_home_education, "2012-13"),
    "Recovered_Sheet17": (parse_home_education, "2011-12"),

    # === NONPUBLIC ENROLLMENT (current + 15 historical) ===
    "NonPublic School Enrollments by": (parse_nonpublic_enrollment, "2025-26"),
    "NonPublic School Enrollments _1": (parse_nonpublic_enrollment, "2025-26"),  # duplicate
    "NonPublic School Enrollments _2": (parse_nonpublic_enrollment, "2024-25"),
    "NonPublic School Enrollments _3": (parse_nonpublic_enrollment, "2023-24"),
    "NonPublic School Enrollments _4": (parse_nonpublic_enrollment, "2022-23"),
    "NonPublic School Enrollments _5": (parse_nonpublic_enrollment, "2021-22"),
    "NonPublic School Enrollments _6": (parse_nonpublic_enrollment, "2020-21"),
    "NonPublic School Enrollments _7": (parse_nonpublic_enrollment, "2019-20"),
    "NonPublic School Enrollments _8": (parse_nonpublic_enrollment, "2018-19"),
    "NonPublic School Enrollments _9": (parse_nonpublic_enrollment, "2017-18"),
    "Recovered_Sheet18": (parse_nonpublic_enrollment, "2016-17"),
    "Recovered_Sheet19": (parse_nonpublic_enrollment, "2015-16"),
    "Recovered_Sheet20": (parse_nonpublic_enrollment, "2014-15"),
    "Recovered_Sheet21": (parse_nonpublic_enrollment, "2013-14"),
    "Recovered_Sheet22": (parse_nonpublic_enrollment, "2012-13"),
    "Recovered_Sheet23": (parse_nonpublic_enrollment, "2011-12"),

    # === OTHER ENROLLMENT ===
    # County Enrollments only has State Total, no per-county breakdown — moved to SKIP_SHEETS
    "Town Level Enrollment By Grade_": (parse_town_enrollment, "2025-26"),
    "Kindergarten Enrollments_Kinder": (parse_kindergarten, "2025-26"),
    "Preschool Enrollments_Preschool": (parse_preschool, "2025-26"),
    "High School Enrollments_High Sc": (parse_hs_enrollment, "2025-26"),

    # === COST PER PUPIL (7 years) ===
    "cost-per-pupil-fy2024-excluding": (parse_cost_per_pupil, "FY2024"),
    "cpp-fy2023_Cost Per Pupil FY202": (parse_cost_per_pupil, "FY2023"),
    "cpp-fy2022_Cost Per Pupil FY202": (parse_cost_per_pupil, "FY2022"),
    "cpp-fy2021_Cost Per Pupil FY202": (parse_cost_per_pupil, "FY2021"),
    "costperpupil2019-20web_Cost Per": (parse_cost_per_pupil, "FY2020"),
    "cost-pupil-district18-19_Cost P": (parse_cost_per_pupil, "FY2019"),
    "cost_per_pupil_by_district_17-1": (parse_cost_per_pupil, "FY2018"),

    # === FREE/REDUCED LUNCH ===
    "Free Reduced K-12 School Lunch ": (parse_free_reduced_lunch, "2025-26"),
    "Free Reduced PRE-School Lunch E": (parse_free_reduced_preschool, "2025-26"),

    # === DEMOGRAPHICS ===
    "Limited English Proficiency Enr": (parse_lep, "2025-26"),
    "Race - Ethnic Enrollments_Race ": (parse_race_ethnic_district, "2025-26"),
    "Race-Ethnic Enrollments by Scho": (parse_race_ethnic_school, "2025-26"),

    # === OUTCOMES ===
    "Attendance Rate By District_Att": (parse_attendance, "2024-25"),
    "Cohort Counts By School_Cohort ": (parse_cohort, "2024-25"),
    "Completers By Category_Complete": (parse_completers_category, "2024-25"),
    # Completers By County only has state total, no per-county breakdown — moved to SKIP_SHEETS
    # "Completers By Status By County_": (parse_completers_county, "2024-25"),
    "Completers By Status By School_": (parse_completers_school, "2024-25"),

    # === STAFFING ===
    "Average Class Size By District_": (parse_class_size_district, "2024-25"),
    "Average Class Size By School_Av": (parse_class_size_school, "2024-25"),
    "Staff Full-time Equivalent by D": (parse_staff_fte, "2024-25"),
    "Student Teacher Ratio by Distri": (parse_student_teacher_ratio, "2024-25"),
    "Teacher Attainment_Teacher Atta": (parse_teacher_attainment, "2024-25"),

    # === TEACHER SALARIES (recent format: 2022-2025) ===
    "teach-sal24-25_Teacher Average ": (parse_teacher_salary, "2024-25"),
    "teach-sal23-24_Teacher Average ": (parse_teacher_salary, "2023-24"),
    "teach-sal22-23_Teacher Average ": (parse_teacher_salary, "2022-23"),

    # === TEACHER SALARIES (older format: 2017-2022) ===
    "teach-sal21-22f_Teacher Average": (parse_teacher_salary_early, "2021-22"),
    "teach-sal20-21_Teacher Average ": (parse_teacher_salary_early, "2020-21"),
    "teach-sal19-20_Publish - Teache": (parse_teacher_salary_early, "2019-20"),
    "teach_sal18-19_Teacher Average ": (parse_teacher_salary_early, "2018-19"),
    "teach_sal17-18_Teacher Average ": (parse_teacher_salary_early, "2017-18"),

    # === TEACHER SALARY SCHEDULES ===
    "teach-sal-sched24-25_Teacher Sa": (parse_teacher_salary_schedule, "2024-25"),
    "teach-sal-schedl23-24_Teacher S": (parse_teacher_salary_schedule, "2023-24"),
    "teach-sal-sched21-22_Publish - ": (parse_teacher_salary_schedule, "2021-22"),
    "teach-sal-sched20-21_Teacher Sa": (parse_teacher_salary_schedule, "2020-21"),
    "teach_sal_sched19-20_Teacher Sa": (parse_teacher_salary_schedule, "2019-20"),
    "teach_sal_sched18-19_Teacher Sa": (parse_teacher_salary_schedule, "2018-19"),

    # === TEACHER STARTING/MIN SALARY ===
    "teach-start-23-24_Teacher Salar": (parse_teacher_salary_schedule, "2023-24"),
    "teach-start-22-23_Teacher Minim": (parse_teacher_starting_salary, "2022-23"),
    "teach-start-21-22_Teacher Minim": (parse_teacher_starting_salary, "2021-22"),
    "teach-start20-21_Teacher Minimu": (parse_teacher_starting_salary, "2020-21"),
    "teach-start19-20_2019-2020": (parse_teacher_starting_salary, "2019-20"),
    "teach_start18-19_Sheet1": (parse_teacher_starting_salary, "2018-19"),

    # === PRINCIPAL SALARIES ===
    "principal-salaries24-25_Sheet1": (parse_principal_salary, "2024-25"),
    "principal-salaries23-24_0_Sheet": (parse_principal_salary, "2023-24"),
    "principal-salaries22-23_0_Sheet": (parse_principal_salary, "2022-23"),
    "principal-salaries21-22_Princip": (parse_principal_salary, "2021-22"),
    "principal-salaries20-21_Publish": (parse_principal_salary, "2020-21"),
    "principal-salaries19-20_Princip": (parse_principal_salary, "2019-20"),

    # === ADMIN SALARIES ===
    "admin-sal-24-25_Sheet1": (parse_admin_salary, "2024-25"),
    "admin-sal-23-24_Sheet3": (parse_admin_salary, "2023-24"),
    "admin-sal-22-23_admin-sal-22-23": (parse_admin_salary, "2022-23"),
    "admin-sal-21-22_SAUAdminSalary": (parse_admin_salary, "2021-22"),
    "salaries20-21_SAUAdminSalary": (parse_admin_salary, "2020-21"),
    "salaries19-20_SAUAdminSalary": (parse_admin_salary, "2019-20"),
    "salaries18-19_Sheet1": (parse_admin_salary_by_name, "2018-19"),

    # === INDIRECT COST RATES ===
    "id-cost-sheet-fy26_INDIRC26 WEB": (parse_indirect_cost, "FY2026"),
    "indirect-cost-fy25_ID FY25": (parse_indirect_cost, "FY2025"),
    "indirect2023-24_INDIRC24": (parse_indirect_cost, "FY2024"),
    "indirect2022-23_INDIRCFY23": (parse_indirect_cost, "FY2023"),
    "indirect2021-22_INDIRC21": (parse_indirect_cost, "FY2022"),
    "indirect2020-21_INDIRC20": (parse_indirect_cost, "FY2021"),
    "indirect19_20_INDIRC20": (parse_indirect_cost, "FY2020"),

    # === EQUALIZED VALUATION PER PUPIL ===
    "eqprt24_Equalized Valuation Per": (parse_equalized_valuation, "2024-25"),
    "eqprt23_Equalized Valuation Per": (parse_equalized_valuation, "2023-24"),
    "eqprt22_Equalized Valuation Per": (parse_equalized_valuation, "2022-23"),
    "eqprt21-_Equalized Valuation Pe": (parse_equalized_valuation, "2021-22"),
    "EVPP-FY-2021-Excel_0_Equalized ": (parse_equalized_valuation, "2020-21"),
    "equal-pupil19-20_Equalized Valu": (parse_equalized_valuation, "2019-20"),
    "equal-pupil18-19_Equalized Valu": (parse_equalized_valuation, "2018-19"),

    # === ESTIMATED EXPENDITURES ===
    "est_exp18-19_Est - Print": (parse_estimated_expenditures, "FY2019"),
    "est_exp19-20_Est - Print": (parse_estimated_expenditures, "FY2020"),
    "est-expend-2020-21_Est - Print": (parse_estimated_expenditures, "FY2021"),
    "est-expend-2021-22_Est - Print": (parse_estimated_expenditures, "FY2022"),
    "est-expend-2022-23_Est. 2023": (parse_estimated_expenditures, "FY2023"),

    # === STATE AVERAGE COST PER PUPIL ===
    "state-avg-cost-per-pupil-fy2024": (parse_state_avg_cpp, "FY2024"),
    "state-avg-cpp-fy2023_State Aver": (parse_state_avg_cpp, "FY2023"),
    "state-avg-cpp-fy2022_State Aver": (parse_state_avg_cpp, "FY2022"),
    "state-avg-cpp-fy2021_State Aver": (parse_state_avg_cpp, "FY2021"),
    "stateavgcostperpupil2019-20web_": (parse_state_avg_cpp, "FY2020"),
    "cost-state-average18-19_State A": (parse_state_avg_cpp, "FY2019"),
    "state_avg_cost_per_pupil17_18_S": (parse_state_avg_cpp, "FY2018"),

    # === STATE SUMMARY REVENUE AND EXPENDITURES ===
    "summary-revexp-fy2024-excluding": (parse_summary_rev_exp, "FY2024"),
    "summary-of-rev-exp-fy2023_State": (parse_summary_rev_exp, "FY2023"),
    "summary-of-rev-exp-fy2022_State": (parse_summary_rev_exp, "FY2022"),
    "summary-of-rev.-exp-fy2021_Stat": (parse_summary_rev_exp, "FY2021"),
    "summaryrevexp-2019-2020web_Stat": (parse_summary_rev_exp, "FY2020"),
    "state-summary18-19_State Summar": (parse_summary_rev_exp, "FY2019"),
    "state_summary_rev_exp17_18_Stat": (parse_summary_rev_exp, "FY2018"),

    # === ADM ===
    "ADM In Attendance and Residence": (parse_adm, "2024-25"),
    "ADM In Attendance and Residen_1": (parse_adm, "2023-24"),
    "ADM In Attendance and Residen_2": (parse_adm, "2022-23"),
    "ADM In Attendance and Residen_3": (parse_adm, "2021-22"),
    "ADM In Attendance and Residen_4": (parse_adm, "2020-21"),
    "ADM In Attendance and Residen_5": (parse_adm, "2019-20"),
    "ADM In Attendance and Residen_6": (parse_adm, "2018-19"),
    "ADM In Attendance and Residen_7": (parse_adm, "2017-18"),
    "ADM In Attendance and Residen_8": (parse_adm, "2016-17"),
    "ADM In Attendance and Residen_9": (parse_adm, "2015-16"),

    # === TRENDS ===
    "Home Education Fall Enrollments": (parse_home_ed_trend, "2016-2026"),
    "Home Education Fall Enrollmen_1": (parse_home_ed_trend, "2016-2026"),
    "State Totals Ten Years Public a": (parse_state_totals, "2016-2026"),
    "State Totals Ten Years Public_1": (parse_state_totals, "2016-2026"),
    "State Totals Ten Years Public_2": (parse_state_totals, "2016-2026"),
    "State Totals Ten Years Public_3": (parse_state_totals, "2016-2026"),

    # === PROPERTY TAX VALUES (7 years) ===
    "values-report-for-web_VALUES 20": (parse_values_tax, "2024"),
    "values-2023_VALUES 2023": (parse_values_tax, "2023"),
    "values-2022_VALUES 2022": (parse_values_tax, "2022"),
    "values-2021_VALUES 2021": (parse_values_tax, "2021"),
    "values-2020-revised_0_VALUES 20": (parse_values_tax, "2020"),
    "equalized-values-2019_VALUES 20": (parse_values_tax, "2019"),
    "values2018_VALUES 2018": (parse_values_tax, "2018"),

    # === STATE PROGRAMS ===
    "25stbyprogram_FY 2023-2025 Stat": (parse_state_programs, "FY2023-2025"),

    # === ASSESSMENT MINIMAL (2024 format) ===
    "minimal-spreadsheet-2024_Sheet1": (parse_assessment_2024_format, "2023-24"),

    # === ASSESSMENT MINIMAL (2023 format) ===
    "minimalspreadsheet_Sheet2": (parse_assessment_2023_format, "2022-23"),

    # === ASSESSMENT MINIMAL (2018-2022 format — existing) ===
    "assessment22-minimal_Sheet1": (parse_assessment_2022_format, "2021-22"),
    "assessment21-minimal_Sheet2": (parse_assessment_2022_format, "2020-21"),
    "assessment19-minimal_Sheet1": (parse_assessment_2022_format, "2018-19"),
    "assessment18-minimal_Sheet1": (parse_assessment_2022_format, "2017-18"),

    # === ASSESSMENT (older format: 2009-2017) ===
    "assessment17_Sheet1": (parse_assessment_old_format, "2016-17"),
    "assessment16_Sheet1": (parse_assessment_old_format, "2015-16"),
    "assessment15_Sheet1": (parse_assessment_old_format, "2014-15"),
    "assessment14_Sheet1": (parse_assessment_old_format, "2013-14"),
    "assessment13_Sheet1": (parse_assessment_old_format, "2012-13"),
    "assessment12_Sheet1": (parse_assessment_old_format, "2011-12"),
    "assessment11_Sheet1": (parse_assessment_old_format, "2010-11"),
    "assessment10_Sheet1": (parse_assessment_old_format, "2009-10"),
    "assessment09_Sheet1": (parse_assessment_old_format, "2008-09"),

    # === DISAGGREGATED ASSESSMENT (large) ===
    "publicdatadisaggregated_sheet1": (parse_assessment_disaggregated, "2024-25"),
    "disagdata2024_Sheet1": (parse_assessment_disaggregated, "2023-24"),
    "disaggregateddata2023_Sheet1": (parse_assessment_disaggregated, "2022-23"),
    "assessment22_Sheet1": (parse_assessment_disaggregated, "2021-22"),
    "assessment21_Sheet2": (parse_assessment_disaggregated, "2020-21"),
    "assessment19_Sheet1": (parse_assessment_disaggregated, "2018-19"),
    "assessment18_Sheet1": (parse_assessment_disaggregated, "2017-18"),

    # === ASSESSMENT BY TOWN (large) ===
    "publicdatabytown_sheet1": (parse_assessment_by_town, "2024-25"),
    "disagdata2024bytownresponsible_": (parse_assessment_by_town, "2023-24"),
    "disagdata2022_2023_2024_bytownr": (parse_assessment_by_town, "2022-2024"),
}

# Sheets we explicitly skip (footnotes, empty, duplicate content, or Manifest)
SKIP_SHEETS = {
    "Manifest",
    "values-report-for-web_Footnotes",
    "values-2022_Footnotes",
    "values-2021_Footnotes",
    "values-2020-revised_0_Footnotes",
    "equalized-values-2019_Footnotes",
    "values2018_Footnotes",
    "assessment18_Sheet2", "assessment18_Sheet3",
    "assessment18-95percent_Sheet1", "assessment18-95percent_Sheet2", "assessment18-95percent_Sheet3",
    "assessment17_Sheet2", "assessment17_Sheet3",
    "assessment16_Sheet2", "assessment16_Sheet3",
    "assessment15_Sheet2", "assessment15_Sheet3",
    "assessment14_Sheet2", "assessment14_Sheet3",
    "assessment13_Sheet2", "assessment13_Sheet3",
    "assessment12_Sheet2", "assessment12_Sheet3",
    "assessment11_Sheet2", "assessment11_Sheet3",
    "assessment10_Sheet2", "assessment10_Sheet3",
    "assessment09_Sheet2", "assessment09_Sheet3",
    "assessment19-95percent_Sheet1",  # duplicate of assessment19
    "25stbystate_Sheet1",  # empty
    "25stbystate_FY 2023-2025 State ",  # state-by-state federal data (not NH-specific)
    "teach-start19-20_Sheet1",  # secondary metadata sheet
    "teach_start18-19_Sheet2",  # secondary metadata sheet
    "teach_sal_sched18-19_Teacher At",  # teacher attainment duplicate
    "teach_sal_sched18-19_Teacher Mi",  # teacher minimum duplicate
    "principal-salaries23-24_0_She_1",  # secondary sheet
    "salaries18-19_Sheet2", "salaries18-19_Sheet3",
    # Recovered sheets for ADM (use main ADM sheets)
    "Recovered_Sheet1", "Recovered_Sheet2", "Recovered_Sheet3",
    "Recovered_Sheet4", "Recovered_Sheet5",
    # Race/ethnic nonpublic (preschool) — different format, minimal data
    "Race-Ethnic Enrollments by Sc_1",
    # County enrollment only has State Total row, no per-county breakdown
    "County Enrollments by Grade_Cou",
    # Completers by county only has state total, no per-county breakdown
    "Completers By Status By County_",
}


# ============================================================================
# MAIN INGESTION
# ============================================================================

def main():
    logger.info("=== Comprehensive iPlatform ingestion (ALL sheets) ===")
    init_db()

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    all_sheets = set(wb.sheetnames)
    logger.info(f"Opened {XLSX_PATH} with {len(all_sheets)} sheets")

    db = SessionLocal()
    try:
        # Clear old iPlatform data
        deleted = db.query(EducationStatistic).delete()
        db.commit()
        if deleted:
            logger.info(f"Cleared {deleted} old education statistics")

        total = 0
        parsed_sheets = set()
        failed_sheets = []

        for sheet_name, (parser, year) in SHEET_MAP.items():
            if sheet_name not in all_sheets:
                logger.warning(f"Sheet '{sheet_name}' not found — skipping")
                continue
            try:
                ws = wb[sheet_name]
                records = parser(ws, year)
                for r in records:
                    db.add(r)
                db.commit()
                logger.info(f"[{year}] {sheet_name}: {len(records)} records")
                total += len(records)
                parsed_sheets.add(sheet_name)
            except Exception as e:
                db.rollback()
                logger.error(f"FAILED: {sheet_name}: {e}")
                failed_sheets.append((sheet_name, str(e)))

        wb.close()

        # Report coverage
        mapped = set(SHEET_MAP.keys()) | SKIP_SHEETS
        unmapped = all_sheets - mapped - parsed_sheets
        if unmapped:
            logger.info(f"Unmapped sheets ({len(unmapped)}): {sorted(unmapped)[:20]}")

        logger.info(f"Total education statistics: {total:,}")
        logger.info(f"Parsed {len(parsed_sheets)} sheets, {len(failed_sheets)} failed")
        if failed_sheets:
            for name, err in failed_sheets:
                logger.error(f"  FAILED: {name}: {err}")

        # Regenerate embeddings (the generate_all_embeddings function handles filtering)
        if HAS_EMBEDDINGS:
            logger.info("Regenerating embeddings...")
            n_embeddings = generate_all_embeddings(db)
            logger.info(f"Total embeddings: {n_embeddings}")
        else:
            logger.warning("Skipping embedding generation (sentence-transformers not available). Run on server to generate embeddings.")

    except Exception as e:
        logger.error(f"Fatal error: {e}", exc_info=True)
    finally:
        db.close()


if __name__ == "__main__":
    main()
