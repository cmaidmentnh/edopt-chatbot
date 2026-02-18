#!/usr/bin/env python3
"""
Ingest NH DOE iPlatform data (XLSX) into the EdOpt chatbot database.
Parses 6 key sheets: district enrollment, home education, cost per pupil,
nonpublic enrollment, free/reduced lunch, and school enrollment.

Usage: python3 ingest_iplatform.py
"""
import json
import logging
from datetime import datetime, timezone

import openpyxl

from models import init_db, SessionLocal, EducationStatistic
from ingest import generate_all_embeddings

XLSX_PATH = "data/iPlatform20251207-clean.xlsx"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
logger = logging.getLogger("ingest_iplatform")


def _int(val):
    """Safely convert to int, handling None, strings with commas/dollars."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        v = int(val)
        return v if v != 0 else None
    s = str(val).strip().replace(",", "").replace("$", "")
    try:
        v = int(float(s))
        return v if v != 0 else None
    except (ValueError, TypeError):
        return None


def _float(val):
    """Safely convert to float."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(",", "").replace("$", "").replace("%", "")
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _str(val):
    """Safely convert to stripped string."""
    if val is None:
        return None
    return str(val).strip()


def parse_district_enrollment(ws):
    """Parse 'District Fall Enrollments_Distr' sheet.
    Header row 12, data from row 14 (row 13 is State Totals).
    Cols: SAU#, SAU Name, District#, District Name, PreSchool, Kindergarten,
          Elementary, Middle, High, PG, Total
    """
    records = []
    for row in ws.iter_rows(min_row=14, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        total = _int(vals[10])
        if total is None:
            continue
        data = {
            "preschool": _int(vals[4]),
            "kindergarten": _int(vals[5]),
            "elementary": _int(vals[6]),
            "middle": _int(vals[7]),
            "high": _int(vals[8]),
            "pg": _int(vals[9]),
            "total": total,
        }
        records.append(EducationStatistic(
            stat_type="district_enrollment",
            school_year="2025-26",
            sau_number=_int(vals[0]),
            sau_name=_str(vals[1]),
            district_number=_int(vals[2]),
            district_name=district_name,
            data_json=json.dumps(data),
        ))
    return records


def parse_home_education(ws):
    """Parse 'Home Education Enrollments By D' sheet.
    Header row 11, data from row 12.
    Cols: SAU#, SAU Name, District#, District Name, Total
    """
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        total = _int(vals[4])
        if total is None:
            continue
        records.append(EducationStatistic(
            stat_type="home_education",
            school_year="2025-26",
            sau_number=_int(vals[0]),
            sau_name=_str(vals[1]),
            district_number=_int(vals[2]),
            district_name=district_name,
            data_json=json.dumps({"total": total}),
        ))
    return records


def parse_cost_per_pupil(ws):
    """Parse 'cost-per-pupil-fy2024-excluding' sheet.
    Header row 19, row 21 is State Average, data from row 23.
    Cols: DIST, LOC, SAU, School District, Elementary, Middle, High, Total(Pre School-12)
    """
    records = []
    for row in ws.iter_rows(min_row=23, values_only=True):
        vals = list(row)
        district_name = _str(vals[3])
        if not district_name:
            continue
        total = _float(vals[7])
        if total is None or total == 0:
            continue
        data = {
            "elementary": _float(vals[4]),
            "middle": _float(vals[5]),
            "high": _float(vals[6]),
            "total": round(total),
        }
        # Round non-None values
        for k in ("elementary", "middle", "high"):
            if data[k] is not None and data[k] != 0:
                data[k] = round(data[k])
            else:
                data[k] = None
        records.append(EducationStatistic(
            stat_type="cost_per_pupil",
            school_year="FY2024",
            sau_number=_int(vals[2]),
            district_number=_int(vals[0]),
            district_name=district_name,
            data_json=json.dumps(data),
        ))
    return records


def parse_nonpublic_enrollment(ws):
    """Parse 'NonPublic School Enrollments by' sheet.
    Header row 10, row 11 is State Total, data from row 12.
    Cols (21): School#, School Name, Town, PS, KG, 1-12, PG, UGEL, UGSE, Total
    """
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        school_name = _str(vals[1])
        if not school_name:
            continue
        total = _int(vals[20])
        if total is None:
            continue
        data = {
            "ps": _int(vals[3]),
            "kg": _int(vals[4]),
        }
        # Grades 1-12
        for g in range(1, 13):
            data[f"grade{g}"] = _int(vals[4 + g])
        data["pg"] = _int(vals[17])
        data["total"] = total
        records.append(EducationStatistic(
            stat_type="nonpublic_enrollment",
            school_year="2025-26",
            school_number=_int(vals[0]),
            school_name=school_name,
            town=_str(vals[2]),
            data_json=json.dumps(data),
        ))
    return records


def parse_free_reduced_lunch(ws):
    """Parse 'Free Reduced K-12 School Lunch ' sheet.
    Header row 10, rows 11-16 are summary/averages, data from row 18.
    Cols: SAU#, SAU Name, District#, District Name, School#, School Name,
          Enroll(1), Free and Reduced Eligible(1), % Eligible
    """
    records = []
    for row in ws.iter_rows(min_row=18, values_only=True):
        vals = list(row)
        school_name = _str(vals[5])
        district_name = _str(vals[3])
        if not school_name and not district_name:
            continue
        # Skip section headers like "District Public Schools"
        if not _int(vals[0]) and not _int(vals[4]):
            continue
        enrollment = _int(vals[6])
        eligible = _int(vals[7])
        pct = _float(vals[8])
        if enrollment is None:
            continue
        data = {
            "enrollment": enrollment,
            "eligible": eligible or 0,
            "pct_eligible": round(pct * 100, 1) if pct is not None and pct < 1 else pct,
        }
        records.append(EducationStatistic(
            stat_type="free_reduced_lunch",
            school_year="2025-26",
            sau_number=_int(vals[0]),
            sau_name=_str(vals[1]),
            district_number=_int(vals[2]),
            district_name=district_name,
            school_number=_int(vals[4]),
            school_name=school_name,
            data_json=json.dumps(data),
        ))
    return records


def parse_school_enrollment(ws):
    """Parse 'School Enrollments by Grade Pub' sheet.
    Header row 10, row 11 is State Total, data from row 12.
    Cols (22): SAU#, SAU Name, District#, District Name, School#, School Name,
               PreSchool, Kindergarten, 1-12, *PG, Total
    """
    records = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        vals = list(row)
        school_name = _str(vals[5])
        if not school_name:
            continue
        total = _int(vals[21])
        if total is None:
            continue
        data = {
            "preschool": _int(vals[6]),
            "kindergarten": _int(vals[7]),
        }
        for g in range(1, 13):
            data[f"grade{g}"] = _int(vals[7 + g])
        data["pg"] = _int(vals[20])
        data["total"] = total
        records.append(EducationStatistic(
            stat_type="school_enrollment",
            school_year="2025-26",
            sau_number=_int(vals[0]),
            sau_name=_str(vals[1]),
            district_number=_int(vals[2]),
            district_name=_str(vals[3]),
            school_number=_int(vals[4]),
            school_name=school_name,
            data_json=json.dumps(data),
        ))
    return records


SUBJECT_NAMES = {"mat": "Math", "rea": "Reading/ELA", "sci": "Science"}

ASSESSMENT_SHEETS = {
    "assessment22-minimal_Sheet1": "2021-22",
    "assessment21-minimal_Sheet2": "2020-21",
    "assessment19-minimal_Sheet1": "2018-19",
    "assessment18-minimal_Sheet1": "2017-18",
}


def parse_assessment_minimal(ws, school_year):
    """Parse a minimal assessment sheet.
    Header row 1, data from row 2.
    Cols: yearid(0), Level of Data(1), Subject(2), DenominatorType(3),
          District(4), Discode(5), School(6), Schcode(7), Grade(8),
          NumberStudents(9), plevel1(10), plevel2(11), plevel3(12), plevel4(13),
          pAboveprof(14), pBelowProf(15), AvgScore(16)
    """
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        school_name = _str(vals[6])
        if not school_name:
            continue
        district_name = _str(vals[4])
        subject = _str(vals[2])
        grade_raw = vals[8]
        grade = "all" if grade_raw in (0, "0") else grade_raw
        above = _str(vals[14])
        below = _str(vals[15])
        avg_score = _str(vals[16])
        data = {
            "subject": subject,
            "subject_name": SUBJECT_NAMES.get(subject, subject),
            "grade": grade,
            "num_students": _str(vals[9]),
            "pct_above_proficient": above,
            "pct_below_proficient": below,
            "avg_score": avg_score if avg_score and avg_score != "Not available" else None,
            "level1_pct": _str(vals[10]),
            "level2_pct": _str(vals[11]),
            "level3_pct": _str(vals[12]),
            "level4_pct": _str(vals[13]),
        }
        records.append(EducationStatistic(
            stat_type="assessment",
            school_year=school_year,
            district_name=district_name,
            school_number=_int(vals[7]),
            school_name=school_name,
            data_json=json.dumps(data),
        ))
    return records


SHEET_PARSERS = {
    "District Fall Enrollments_Distr": parse_district_enrollment,
    "Home Education Enrollments By D": parse_home_education,
    "cost-per-pupil-fy2024-excluding": parse_cost_per_pupil,
    "NonPublic School Enrollments by": parse_nonpublic_enrollment,
    "Free Reduced K-12 School Lunch ": parse_free_reduced_lunch,
    "School Enrollments by Grade Pub": parse_school_enrollment,
}


def main():
    logger.info("Ingesting NH DOE iPlatform data...")
    init_db()

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    logger.info(f"Opened {XLSX_PATH} with {len(wb.sheetnames)} sheets")

    db = SessionLocal()
    try:
        # Clear old iPlatform data
        deleted = db.query(EducationStatistic).delete()
        db.commit()
        if deleted:
            logger.info(f"Cleared {deleted} old education statistics")

        total = 0
        for sheet_name, parser in SHEET_PARSERS.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                records = parser(ws)
                for r in records:
                    db.add(r)
                db.commit()
                logger.info(f"Stored {len(records)} records from '{sheet_name}'")
                total += len(records)
            else:
                logger.warning(f"Sheet '{sheet_name}' not found — skipping")

        # Assessment sheets (need year parameter)
        for sheet_name, year in ASSESSMENT_SHEETS.items():
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                records = parse_assessment_minimal(ws, year)
                for r in records:
                    db.add(r)
                db.commit()
                logger.info(f"Stored {len(records)} records from '{sheet_name}' ({year})")
                total += len(records)
            else:
                logger.warning(f"Sheet '{sheet_name}' not found — skipping")

        wb.close()
        logger.info(f"Total education statistics: {total}")

        # Regenerate all embeddings (includes new education stats)
        n_embeddings = generate_all_embeddings(db)
        logger.info(f"Total embeddings: {n_embeddings}")

    except Exception as e:
        logger.error(f"Failed: {e}", exc_info=True)
    finally:
        db.close()


if __name__ == "__main__":
    main()
